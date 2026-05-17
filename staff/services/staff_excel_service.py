import openpyxl
from django.http import HttpResponse
from ..models import Staff

# -------------------------------------------------------------------
# FIELD ORDER (EXACT ORDER OF MODEL FIELDS)
# -------------------------------------------------------------------

STAFF_EXCEL_FIELDS = [
    "staff_id",
    "name",
    "program",
    "department",
    "designation",
    "college",
    "doj",
    "dor",
    "city",
    "district",
    "email",
    "phone",
    "bank_account",
    "bank_name",
    "ifsc_code",
    "remark",
]

# -------------------------------------------------------------------
# SAMPLE FILE
# -------------------------------------------------------------------

def download_sample_staff_excel():

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(STAFF_EXCEL_FIELDS)
    ws.append([
        "JM001", "Karunya", "UG", "CS", "Professor", "ABC College", "2020-01-01", 
        "2025-01-01", "Chennai", "Chennai", "test@mail.com", "9876543210", 
        "1234567890", "SBI", "SBIN0001234", "Sample remark"
    ])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Staff_Template.xlsx"'
    wb.save(response)
    return response

# -------------------------------------------------------------------
# EXPORT (MODEL → EXCEL)
# -------------------------------------------------------------------

def export_staff_excel():

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(STAFF_EXCEL_FIELDS)
    for s in Staff.objects.all():
       ws.append([
            s.staff_id, s.name, s.program, s.department, s.designation, s.college,
            s.doj, s.dor, s.city, s.district, s.email, s.phone,
            s.bank_account, s.bank_name, s.ifsc_code, s.remark,
        ])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Staff_Data.xlsx"'
    wb.save(response)
    return response

# -------------------------------------------------------------------
# UPLOAD 
# -------------------------------------------------------------------

def upload_staff_excel(file):

    wb = openpyxl.load_workbook(file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [h.strip() for h in rows[0]]
    for row in rows[1:]:
        data = dict(zip(headers, row))
        Staff.objects.create(
            staff_id=data.get("staff_id"),
            name=data.get("name"),
            program=data.get("program"),
            department=data.get("department"),
            designation=data.get("designation"),
            college=data.get("college"),
            doj=data.get("doj"),
            dor=data.get("dor"),
            city=data.get("city"),
            district=data.get("district"),
            email=data.get("email"),
            phone=data.get("phone"),
            bank_account=data.get("bank_account"),
            bank_name=data.get("bank_name"),
            ifsc_code=data.get("ifsc_code"),
            remark=data.get("remark"),
        )

# -------------------------------------------------------------------