from django.shortcuts import render, get_object_or_404
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from .models import Staff
from core.models import FieldReference, AuditLog, SystemSetting
import json
import logging

logger = logging.getLogger(__name__)

# -------------------------------------------------------------------------------------------------------------------------

@login_required
def staff_list(request):

    staff_queryset = Staff.objects.all()
    ALLOWED_PER_PAGE = [25, 50, 75, 100, 500, 1000]
    default_per_page = SystemSetting.get_setting('staff_items_per_page', 500)
    per_page = request.GET.get('per_page', default_per_page)

    try:
        per_page = int(per_page)
        if per_page not in ALLOWED_PER_PAGE:
            per_page = default_per_page
    except (ValueError, TypeError):
        per_page = default_per_page

    # --- Sorting ---
    sort_by = request.GET.get('sort', 'id')
    sort_dir = request.GET.get('dir', 'desc')

    valid_sort_fields = [
        'id', 'staff_id', 'name', 'designation', 'program',
        'college', 'district', 'doj', 'dor', 'phone', 'email',
        'bank_account', 'bank_name', 'ifsc_code', 'created_at'
    ]

    if sort_by in valid_sort_fields:
        if sort_dir == 'desc':
            staff_queryset = staff_queryset.order_by(f'-{sort_by}')
        else:
            staff_queryset = staff_queryset.order_by(sort_by)
    else:
        staff_queryset = staff_queryset.order_by('-id')

    # --- Search ---
    search_query = request.GET.get('search', '')
    if search_query:
        staff_queryset = staff_queryset.filter(
            Q(id__icontains=search_query) |
            Q(staff_id__icontains=search_query) |
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(ifsc_code__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(designation__icontains=search_query) |
            Q(program__icontains=search_query) |
            Q(college__icontains=search_query) |
            Q(district__icontains=search_query) |
            Q(bank_name__icontains=search_query) |
            Q(staff_category__icontains=search_query) |
            Q(dept_category__icontains=search_query) |
            Q(branch__icontains=search_query)
        )

    # --- Filters ---
    filters = {
        'designation': request.GET.get('designation', ''),
        'staff_category': request.GET.get('staff_category', ''),
        'dept_category': request.GET.get('dept_category', ''),
        'branch': request.GET.get('branch', ''),
        'program': request.GET.get('program', ''),
        'college': request.GET.get('college', ''),
    }

    for field, value in filters.items():
        if value:
            filter_kwargs = {field: value}
            staff_queryset = staff_queryset.filter(**filter_kwargs)

    filter_options = {
        'designations': FieldReference.get_options(Staff, 'designation'),
        'staff_categories': FieldReference.get_options(Staff, 'staff_category'),
        'dept_categories': FieldReference.get_options(Staff, 'dept_category'),
        'branches': FieldReference.get_options(Staff, 'branch'),
        'programs': FieldReference.get_options(Staff, 'program'),
        'colleges': FieldReference.get_options(Staff, 'college'),
    }

    paginator = Paginator(staff_queryset, per_page)
    page_number = request.GET.get('page', 1)
    staff_page = paginator.get_page(page_number)

    if SystemSetting.get_setting('log_list_views', False):
        AuditLog.log(
            request=request,
            action='VIEW',
            obj=None,
            object_repr=f"Staff list viewed - Page {page_number}",
            changes={
                'search': search_query,
                'filters': filters,
                'total_count': staff_queryset.count()
            }
        )

    context = {
        'staff': staff_page,
        'total_count': staff_queryset.count(),
        'search': search_query,
        'filters': filters,
        'filter_options': filter_options,
        'current_sort': sort_by,
        'current_dir': sort_dir,
        'per_page': per_page,                     
        'allowed_per_page': ALLOWED_PER_PAGE,    
    }

    return render(request, 'staff/staff_list.html', context)
    
# -------------------------------------------------------------------------------------------------------------------------

@require_http_methods(["GET"])
def get_field_options(request):
  
    field_name = request.GET.get('field')
    search_query = request.GET.get('search', '')
    
    if not field_name:
        return JsonResponse({'success': False, 'error': 'Field name is required'}, status=400)
    
    # Map frontend field names to actual model field names - removed program_type
    field_mapping = {
        'designation': 'designation',
        'staff_category': 'staff_category',
        'dept_category': 'dept_category',
        'examiner_type': 'examiner_type',
        'branch': 'branch',
        'branch_final': 'branch_final',
        'program': 'program',
        'college': 'college',
        'qualification': 'qualification',
        'district': 'district',
        'bank_name': 'bank_name',
        'bank_city': 'bank_city',
        'place': 'place',
    }
    
    model_field = field_mapping.get(field_name)
    if not model_field:
        return JsonResponse({'success': False, 'error': f'Invalid field name: {field_name}'}, status=400)
    
    try:
        options = FieldReference.get_options(
            model=Staff,
            field_name=model_field,
            search=search_query
        )
        
        return JsonResponse({
            'success': True,
            'options': options,
            'total_count': len(options)
        })
        
    except Exception as e:
        logger.error(f"Error getting field options: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# -------------------------------------------------------------------------------------------------------------------------

@require_http_methods(["GET", "POST"])
@login_required
def staff_add(request):
   
    if request.method == "POST":
        try:
            # Handle both JSON and form data
            if request.content_type == 'application/json':
                data = json.loads(request.body)
            else:
                data = request.POST

            # Check if phone already exists
            phone = data.get('phone')
            if phone and Staff.objects.filter(phone=phone).exists():
                return JsonResponse({'success': False, 'error': 'Phone number already exists'}, status=400)
            
            # Check if staff_id already exists
            staff_id = data.get('staff_id')
            if staff_id and Staff.objects.filter(staff_id=staff_id).exists():
                return JsonResponse({'success': False, 'error': 'Staff ID already exists'}, status=400)

            # Create staff member
            staff = Staff()

            # Map form fields to model fields - added password, removed program_type
            fields_mapping = {
                'staff_id': 'staff_id',
                'name': 'name',
                'designation': 'designation',
                'program': 'program',
                'college': 'college',
                'doj': 'doj',
                'dor': 'dor',
                'phone': 'phone',
                'password': 'password',
                'email': 'email',
                'district': 'district',
                'bank_account': 'bank_account',
                'bank_name': 'bank_name',
                'ifsc_code': 'ifsc_code',
                'remark': 'remark',
                'staff_category': 'staff_category',
                'dept_category': 'dept_category',
                'examiner_type': 'examiner_type',
                'branch': 'branch',
                'branch_final': 'branch_final',
                'place': 'place',
                'qualification': 'qualification',
                'bank_city': 'bank_city',
                'branch_code': 'branch_code',
            }

            changes = {}
            
            for form_field, model_field in fields_mapping.items():
                value = data.get(form_field)
                if value and str(value).strip():
                    # Handle date fields
                    if model_field in ['doj', 'dor'] and value:
                        try:
                            date_value = datetime.strptime(str(value), '%Y-%m-%d').date()
                            setattr(staff, model_field, date_value)
                            changes[model_field] = str(date_value)
                        except:
                            setattr(staff, model_field, None)
                    else:
                        str_value = str(value).strip()
                        setattr(staff, model_field, str_value)
                        changes[model_field] = str_value
                else:
                    setattr(staff, model_field, None)

            staff.full_clean()
            staff.save()
            
            AuditLog.log(
                request=request,
                action='CREATE',
                obj=staff,
                object_repr=f"{staff.staff_id} - {staff.name}",
                changes=changes
            )
            
            return JsonResponse({
                'success': True, 
                'message': 'Staff added successfully', 
                'id': staff.id
            })

        except ValidationError as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        except Exception as e:
            logger.error(f"Error adding staff: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

    # GET request - return form with options from FieldReference
    context = {
        'field_options': {
            'designations': FieldReference.get_options(Staff, 'designation'),
            'staff_categories': FieldReference.get_options(Staff, 'staff_category'),
            'dept_categories': FieldReference.get_options(Staff, 'dept_category'),
            'examiner_types': FieldReference.get_options(Staff, 'examiner_type'),
            'branches': FieldReference.get_options(Staff, 'branch'),
            'branch_finals': FieldReference.get_options(Staff, 'branch_final'),
            'programs': FieldReference.get_options(Staff, 'program'),
            'colleges': FieldReference.get_options(Staff, 'college'),
            'qualifications': FieldReference.get_options(Staff, 'qualification'),
            'districts': FieldReference.get_options(Staff, 'district'),
            'bank_names': FieldReference.get_options(Staff, 'bank_name'),
            'bank_cities': FieldReference.get_options(Staff, 'bank_city'),
        }
    }
    return render(request, 'staff/staff_form.html', context)

# -------------------------------------------------------------------------------------------------------------------------

@require_http_methods(["GET", "POST"])
@login_required
def staff_edit(request, phone):
   
    staff = get_object_or_404(Staff, phone=phone)

    if request.method == "GET":
        data = {
            'success': True,
            'staff': {
                'id': staff.id,
                'staff_id': staff.staff_id,
                'name': staff.name,
                'designation': staff.designation or '',
                'program': staff.program or '',
                'college': staff.college or '',
                'doj': staff.doj.strftime('%Y-%m-%d') if staff.doj else '',
                'dor': staff.dor.strftime('%Y-%m-%d') if staff.dor else '',
                'phone': staff.phone,
                'password': staff.password or '',  # added
                'email': staff.email or '',
                'district': staff.district or '',
                'bank_account': staff.bank_account or '',
                'bank_name': staff.bank_name or '',
                'ifsc_code': staff.ifsc_code or '',
                'remark': staff.remark or '',
                'staff_category': staff.staff_category or '',
                'dept_category': staff.dept_category or '',
                'examiner_type': staff.examiner_type or '',
                'branch': staff.branch or '',
                'branch_final': staff.branch_final or '',
                'place': staff.place or '',
                'qualification': staff.qualification or '',
                'bank_city': staff.bank_city or '',
                'branch_code': staff.branch_code or '',
            }
        }
        return JsonResponse(data)

    elif request.method == "POST":
        try:
            old_values = {
                'staff_id': staff.staff_id,
                'name': staff.name,
                'designation': staff.designation,
                'program': staff.program,
                'college': staff.college,
                'phone': staff.phone,
                'password': staff.password,
                'email': staff.email,
                'district': staff.district,
                'bank_account': staff.bank_account,
                'bank_name': staff.bank_name,
                'ifsc_code': staff.ifsc_code,
                'staff_category': staff.staff_category,
                'dept_category': staff.dept_category,
                'examiner_type': staff.examiner_type,
                'branch': staff.branch,
                'branch_final': staff.branch_final,
                'place': staff.place,
                'qualification': staff.qualification,
                'bank_city': staff.bank_city,
                'branch_code': staff.branch_code,
            }
            
            fields = [
                'staff_id', 'name', 'designation', 'email', 'program',
                'college', 'doj', 'dor', 'district',
                'bank_account', 'bank_name', 'ifsc_code', 'remark',
                'staff_category', 'dept_category', 'examiner_type',
                'branch', 'branch_final', 'place', 'qualification',
                'bank_city', 'branch_code', 'password'  # added password
            ]

            changes = {}
            
            for field in fields:
                value = request.POST.get(field)
                if value and value.strip():
                    if field in ['doj', 'dor'] and value:
                        try:
                            date_value = datetime.strptime(str(value), '%Y-%m-%d').date()
                            setattr(staff, field, date_value)
                            if str(old_values.get(field)) != str(date_value):
                                changes[field] = {'old': str(old_values.get(field)), 'new': str(date_value)}
                        except:
                            setattr(staff, field, None)
                    else:
                        str_value = str(value).strip()
                        setattr(staff, field, str_value)
                        if str(old_values.get(field)) != str_value:
                            changes[field] = {'old': str(old_values.get(field)), 'new': str_value}
                else:
                    setattr(staff, field, None)
                    if old_values.get(field) is not None:
                        changes[field] = {'old': str(old_values.get(field)), 'new': None}

            # Phone check
            new_phone = request.POST.get('phone')
            if new_phone and new_phone.strip():
                new_phone = new_phone.strip()
                if new_phone != old_values['phone']:
                    if Staff.objects.filter(phone=new_phone).exists():
                        return JsonResponse({'success': False, 'error': 'Phone number already exists'}, status=400)
                    staff.phone = new_phone
                    changes['phone'] = {'old': old_values['phone'], 'new': new_phone}
            else:
                return JsonResponse({'success': False, 'error': 'Phone number is required'}, status=400)
            
            # Staff ID check
            new_staff_id = request.POST.get('staff_id')
            if new_staff_id and new_staff_id.strip():
                new_staff_id = new_staff_id.strip()
                if new_staff_id != old_values['staff_id']:
                    if Staff.objects.filter(staff_id=new_staff_id).exists():
                        return JsonResponse({'success': False, 'error': 'Staff ID already exists'}, status=400)
                    staff.staff_id = new_staff_id
                    changes['staff_id'] = {'old': old_values['staff_id'], 'new': new_staff_id}
            else:
                return JsonResponse({'success': False, 'error': 'Staff ID is required'}, status=400)

            staff.full_clean()
            staff.save()
            
            if changes:
                AuditLog.log(
                    request=request,
                    action='UPDATE',
                    obj=staff,
                    object_repr=f"{staff.staff_id} - {staff.name}",
                    changes=changes
                )
            
            return JsonResponse({'success': True, 'message': 'Staff updated successfully'})

        except ValidationError as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        except Exception as e:
            logger.error(f"Error editing staff: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

# -------------------------------------------------------------------------------------------------------------------------

@require_http_methods(["DELETE", "POST"])
@login_required
def staff_delete(request, phone):
    
    try:
        staff = get_object_or_404(Staff, phone=phone)
        
        staff_info = f"{staff.staff_id} - {staff.name}"
        staff_data = {
            'staff_id': staff.staff_id,
            'name': staff.name,
            'phone': staff.phone,
            'email': staff.email,
            'designation': staff.designation
        }
        
        soft_delete = SystemSetting.get_setting('soft_delete_staff', False)
        
        if soft_delete:
            staff.is_active = False
            staff.save()
            message = 'Staff member soft deleted successfully'
            action = 'UPDATE' 
            
            AuditLog.log(
                request=request,
                action=action,
                obj=staff,
                changes=staff_data,
                object_repr=f"Soft deleted: {staff_info}"
            )
        else:
            staff_id = str(staff.id) 
            staff_app_label = staff._meta.app_label
            staff_model_name = staff.__class__.__name__
            
            staff.delete()
            message = 'Staff member deleted successfully'
            action = 'DELETE'
            
            audit_data = {
                'action': action,
                'changes': staff_data,
                'app_label': staff_app_label,
                'model_name': staff_model_name,
                'object_id': staff_id,
                'object_repr': staff_info,
                'ip_address': AuditLog._get_client_ip(request),
                'user_agent': request.META.get('HTTP_USER_AGENT', ''),
                'request_path': request.path,
            }
            
            if request.user.is_authenticated:
                audit_data['user_id'] = str(request.user.id) 
                audit_data['user_name'] = request.user.username  
            
            AuditLog.objects.create(**audit_data)
        
        return JsonResponse({'success': True, 'message': message})
        
    except Exception as e:
        logger.error(f"Error deleting staff: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

# -------------------------------------------------------------------------------------------------------------------------

@require_http_methods(["POST"])
@login_required
def save_field_option(request):

    try:
        data = json.loads(request.body)
        field_name = data.get('field_name')
        value = data.get('value')
        
        if not field_name or not value:
            return JsonResponse({'success': False, 'error': 'Field name and value required'}, status=400)
        
        valid_fields = [
            'designation', 'program', 'college', 'district',
            'bank_name', 'staff_category', 'dept_category', 'examiner_type',
            'branch', 'branch_final', 'place', 'qualification', 'bank_city'
        ]
        
        if field_name not in valid_fields:
            return JsonResponse({'success': False, 'error': 'Invalid field name'}, status=400)
        
        obj, created = FieldReference.add_option(
            model=Staff,
            field_name=field_name,
            value=value,
            created_by=request.user.username
        )
        
        return JsonResponse({
            'success': True,
            'created': created,
            'message': 'Option saved successfully'
        })
        
    except Exception as e:
        logger.error(f"Error saving field option: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

# -------------------------------------------------------------------------------------------------------------------------

@require_http_methods(["POST"])
@login_required
def staff_upload(request):
    if request.method == "POST" and request.FILES.get('file'):
        try:
            file = request.FILES['file']
            if not file.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({'success': False, 'error': 'Invalid file format. Please upload Excel file.'}, status=400)

            wb = load_workbook(file)
            ws = wb.active

            print(f"\n{'='*60}")
            print(f"STARTING STAFF UPLOAD: {file.name}")
            print(f"{'='*60}")

            success_count = 0
            update_count = 0
            error_count = 0
            errors = []
            created_staff_ids = []
            updated_staff_ids = []
            rows_details = []   # for the report

            new_field_options = {
                'designation': set(), 'program': set(), 'college': set(),
                'district': set(), 'bank_name': set(), 'staff_category': set(),
                'dept_category': set(), 'examiner_type': set(), 'branch': set(),
                'branch_final': set(), 'place': set(), 'qualification': set(),
                'bank_city': set(),
            }

            total_rows = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row or not any(row):
                        continue
                    total_rows += 1

                    # Extract values (same as before)
                    program = str(row[0]).strip() if row[0] else None
                    staff_id = str(row[1]).strip() if row[1] else None
                    staff_category = str(row[2]).strip() if row[2] else None
                    dept_category = str(row[3]).strip() if row[3] else None
                    examiner_type = str(row[4]).strip() if row[4] else None
                    name = str(row[5]).strip() if row[5] else None
                    designation = str(row[6]).strip() if row[6] else None
                    branch = str(row[7]).strip() if row[7] else None
                    branch_final = str(row[8]).strip() if row[8] else None
                    college = str(row[9]).strip() if row[9] else None
                    place = str(row[10]).strip() if row[10] else None
                    district = str(row[11]).strip() if row[11] else None
                    qualification = str(row[12]).strip() if row[12] else None

                    doj = None
                    if len(row) > 13 and row[13]:
                        try:
                            doj = datetime.strptime(str(row[13]), '%d-%m-%Y').date()
                        except ValueError:
                            if isinstance(row[13], datetime):
                                doj = row[13].date()
                            else:
                                raise ValueError(f"Invalid DOJ format, expected dd-mm-yyyy: {row[13]}")

                    dor = None
                    if len(row) > 14 and row[14]:
                        try:
                            dor = datetime.strptime(str(row[14]), '%d-%m-%Y').date()
                        except ValueError:
                            if isinstance(row[14], datetime):
                                dor = row[14].date()
                            else:
                                raise ValueError(f"Invalid DOR format, expected dd-mm-yyyy: {row[14]}")

                    phone = str(row[15]).strip() if len(row) > 15 and row[15] else None
                    email = str(row[16]).strip() if len(row) > 16 and row[16] else None
                    bank_account = str(row[17]).strip() if len(row) > 17 and row[17] else None
                    bank_name = str(row[18]).strip() if len(row) > 18 and row[18] else None
                    bank_city = str(row[19]).strip() if len(row) > 19 and row[19] else None
                    branch_code = str(row[20]).strip() if len(row) > 20 and row[20] else None
                    ifsc_code = str(row[21]).strip() if len(row) > 21 and row[21] else None
                    remark = str(row[22]).strip() if len(row) > 22 and row[22] else None
                    password = str(row[23]).strip() if len(row) > 23 and row[23] else ''

                    # ----- Only Phone and Name are required (staff_id is nullable) -----
                    if not phone:
                        msg = "Phone number is required"
                        errors.append(f"Row {row_idx}: {msg}")
                        error_count += 1
                        print(f"  ❌ ERROR Row {row_idx}: {msg}")   # <-- print error
                        rows_details.append({'row': row_idx, 'phone': '', 'staff_id': staff_id or '', 'name': name or '', 'status': 'Error', 'error': msg})
                        continue
                    if not name:
                        msg = "Name is required"
                        errors.append(f"Row {row_idx}: {msg}")
                        error_count += 1
                        print(f"  ❌ ERROR Row {row_idx}: {msg}")   # <-- print error
                        rows_details.append({'row': row_idx, 'phone': phone, 'staff_id': staff_id or '', 'name': '', 'status': 'Error', 'error': msg})
                        continue

                    # Collect field options
                    if designation: new_field_options['designation'].add(designation)
                    if program: new_field_options['program'].add(program)
                    if college: new_field_options['college'].add(college)
                    if district: new_field_options['district'].add(district)
                    if bank_name: new_field_options['bank_name'].add(bank_name)
                    if staff_category: new_field_options['staff_category'].add(staff_category)
                    if dept_category: new_field_options['dept_category'].add(dept_category)
                    if examiner_type: new_field_options['examiner_type'].add(examiner_type)
                    if branch: new_field_options['branch'].add(branch)
                    if branch_final: new_field_options['branch_final'].add(branch_final)
                    if place: new_field_options['place'].add(place)
                    if qualification: new_field_options['qualification'].add(qualification)
                    if bank_city: new_field_options['bank_city'].add(bank_city)

                    staff_data = {
                        'staff_id': staff_id,   # may be None
                        'name': name,
                        'designation': designation,
                        'program': program,
                        'college': college,
                        'phone': phone,
                        'password': password,
                        'email': email,
                        'doj': doj,
                        'dor': dor,
                        'district': district,
                        'bank_account': bank_account,
                        'bank_name': bank_name,
                        'ifsc_code': ifsc_code,
                        'remark': remark,
                        'staff_category': staff_category,
                        'dept_category': dept_category,
                        'examiner_type': examiner_type,
                        'branch': branch,
                        'branch_final': branch_final,
                        'place': place,
                        'qualification': qualification,
                        'bank_city': bank_city,
                        'branch_code': branch_code,
                    }

                    existing_staff = Staff.objects.filter(phone=phone).first()
                    if existing_staff:
                        # Update – but DO NOT overwrite staff_id with None
                        for key, value in staff_data.items():
                            if value is not None or key != 'staff_id':
                                setattr(existing_staff, key, value)
                        try:
                            existing_staff.save()
                            update_count += 1
                            updated_staff_ids.append(existing_staff.id)
                            # No success print – silent
                            rows_details.append({'row': row_idx, 'phone': phone, 'staff_id': staff_id or '', 'name': name, 'status': 'Updated', 'error': ''})
                        except Exception as e:
                            msg = f"Update failed: {str(e)}"
                            errors.append(f"Row {row_idx}: {msg}")
                            error_count += 1
                            print(f"  ❌ ERROR Row {row_idx}: {msg}")   # <-- print error
                            rows_details.append({'row': row_idx, 'phone': phone, 'staff_id': staff_id or '', 'name': name, 'status': 'Error', 'error': msg})
                    else:
                        staff = Staff(**staff_data)
                        try:
                            staff.full_clean()
                            staff.save()
                            success_count += 1
                            created_staff_ids.append(staff.id)
                            # No success print – silent
                            rows_details.append({'row': row_idx, 'phone': phone, 'staff_id': staff_id or '', 'name': name, 'status': 'Created', 'error': ''})
                        except Exception as e:
                            msg = str(e)
                            errors.append(f"Row {row_idx}: {msg}")
                            error_count += 1
                            print(f"  ❌ ERROR Row {row_idx}: {msg}")   # <-- print error
                            rows_details.append({'row': row_idx, 'phone': phone, 'staff_id': staff_id or '', 'name': name, 'status': 'Error', 'error': msg})

                except Exception as e:
                    error_count += 1
                    msg = str(e)
                    errors.append(f"Row {row_idx}: {msg}")
                    print(f"  ❌ ERROR Row {row_idx}: {msg}")   # <-- print error
                    phone = str(row[15]).strip() if len(row) > 15 and row[15] else ''
                    staff_id = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    name = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                    rows_details.append({'row': row_idx, 'phone': phone, 'staff_id': staff_id, 'name': name, 'status': 'Error', 'error': msg})
                    logger.error(f"Error uploading row {row_idx}: {msg}")

            # Store report in session
            request.session['upload_report'] = rows_details

            # ----- Summary (always printed) -----
            print(f"\n{'='*60}")
            print(f"UPLOAD SUMMARY for {file.name}")
            print(f"{'='*60}")
            print(f"Total rows processed: {total_rows}")
            print(f"  ✅ New staff created: {success_count}")
            print(f"  🔄 Existing staff updated: {update_count}")
            print(f"  ❌ Errors: {error_count}")
            if errors:
                print(f"\nFirst 10 errors (out of {len(errors)}):")
                for err in errors[:10]:
                    print(f"  - {err}")
            print(f"{'='*60}\n")

            # Add dropdown options 
            added_options_count = 0
            for field_name, values in new_field_options.items():
                for value in values:
                    if value:
                        try:
                            obj, created = FieldReference.add_option(
                                model=Staff,
                                field_name=field_name,
                                value=value,
                                created_by=request.user.username if request.user.is_authenticated else 'excel_upload'
                            )
                            if created:
                                added_options_count += 1
                        except Exception as e:
                            logger.error(f"Error adding {field_name}={value}: {str(e)}")

            # Audit log 
            AuditLog.objects.create(
                action='UPLOAD',
                app_label=Staff._meta.app_label,
                model_name=Staff.__name__,
                object_repr=f"Bulk upload: {success_count} new, {update_count} updated",
                changes={
                    'filename': file.name,
                    'success_count': success_count,
                    'update_count': update_count,
                    'error_count': error_count,
                    'new_options_added': added_options_count,
                    'created_staff_ids': created_staff_ids[:10],
                    'updated_staff_ids': updated_staff_ids[:10]
                },
                ip_address=AuditLog._get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                request_path=request.path,
                user_id=request.user.id if request.user.is_authenticated else None,
                user_name=request.user.username if request.user.is_authenticated else ''
            )

            # Response (unchanged)
            message_parts = []
            if success_count > 0:
                message_parts.append(f"{success_count} new staff added")
            if update_count > 0:
                message_parts.append(f"{update_count} staff updated")
            if added_options_count > 0:
                message_parts.append(f"{added_options_count} new dropdown options added")

            message = f"Successfully processed: {', '.join(message_parts)}"
            if error_count > 0:
                message += f". {error_count} errors occurred."
                messages.warning(request, message)
            else:
                messages.success(request, message)

            return JsonResponse({
                'success': True,
                'success_count': success_count,
                'update_count': update_count,
                'error_count': error_count,
                'errors': errors[:20],
                'new_options_added': added_options_count,
                'created_staff_ids': created_staff_ids,
                'updated_staff_ids': updated_staff_ids,
                'report_available': True
            })

        except Exception as e:
            logger.error(f"Error in staff upload: {str(e)}")
            print(f"FATAL ERROR during upload: {str(e)}")
            import traceback
            print(traceback.format_exc())
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({'success': False, 'error': 'No file provided'}, status=400)

# -------------------------------------------------------------------------------------------------------------------------

@login_required
def staff_export(request):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Staff Data"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # New header order (matching the required sequence)
        headers = [
            'Program', 'Staff ID', 'Staff Category', 'Department Category',
            'Internal/External Examiner', 'Name', 'Designation', 'Branch',
            'Branch Final', 'College Name', 'Place', 'District Name',
            'Qualification', 'Date of Joining (DOJ)', 'Date of Retirement (DOR)',
            'Phone Number', 'Email ID', 'S.B. Account Number', 'Bank Name',
            'Bank City Name', 'Branch Code', 'IFSC Code', 'Remarks', 'Password'
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        staff_members = Staff.objects.all().order_by('id')
        total_count = staff_members.count()

        for row_idx, staff in enumerate(staff_members, start=2):
            # Write in the same order as headers
            ws.cell(row=row_idx, column=1, value=staff.program or '')
            ws.cell(row=row_idx, column=2, value=staff.staff_id)
            ws.cell(row=row_idx, column=3, value=staff.staff_category or '')
            ws.cell(row=row_idx, column=4, value=staff.dept_category or '')
            ws.cell(row=row_idx, column=5, value=staff.examiner_type or '')
            ws.cell(row=row_idx, column=6, value=staff.name)
            ws.cell(row=row_idx, column=7, value=staff.designation or '')
            ws.cell(row=row_idx, column=8, value=staff.branch or '')
            ws.cell(row=row_idx, column=9, value=staff.branch_final or '')
            ws.cell(row=row_idx, column=10, value=staff.college or '')
            ws.cell(row=row_idx, column=11, value=staff.place or '')
            ws.cell(row=row_idx, column=12, value=staff.district or '')
            ws.cell(row=row_idx, column=13, value=staff.qualification or '')
            ws.cell(row=row_idx, column=14, value=staff.doj.strftime('%d-%m-%Y') if staff.doj else '')
            ws.cell(row=row_idx, column=15, value=staff.dor.strftime('%d-%m-%Y') if staff.dor else '')
            ws.cell(row=row_idx, column=16, value=staff.phone or '')
            ws.cell(row=row_idx, column=17, value=staff.email or '')
            ws.cell(row=row_idx, column=18, value=staff.bank_account or '')
            ws.cell(row=row_idx, column=19, value=staff.bank_name or '')
            ws.cell(row=row_idx, column=20, value=staff.bank_city or '')
            ws.cell(row=row_idx, column=21, value=staff.branch_code or '')
            ws.cell(row=row_idx, column=22, value=staff.ifsc_code or '')
            ws.cell(row=row_idx, column=23, value=staff.remark or '')
            ws.cell(row=row_idx, column=24, value=staff.password or '')

        # Auto-adjust column widths (unchanged)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        AuditLog.objects.create(
            action='EXPORT',
            app_label=Staff._meta.app_label,
            model_name=Staff.__name__,
            object_repr=f"Staff export - {total_count} records",
            changes={
                'total_records': total_count,
                'file_format': 'xlsx',
                'export_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            },
            ip_address=AuditLog._get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            request_path=request.path,
            user_id=request.user.id if request.user.is_authenticated else None,
            user_name=request.user.username if request.user.is_authenticated else ''
        )

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Staff Data.xlsx'
        wb.save(response)
        return response

    except Exception as e:
        logger.error(f"Error exporting staff: {str(e)}")
        messages.error(request, f"Error exporting data: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# -------------------------------------------------------------------------------------------------------------------------

@login_required
def staff_sample(request):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Staff Template"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Same header order as export
        headers = [
            'Program', 'Staff ID', 'Staff Category', 'Department Category',
            'Internal/External Examiner', 'Name', 'Designation', 'Branch',
            'Branch Final', 'College Name', 'Place', 'District Name',
            'Qualification', 'Date of Joining (DOJ)', 'Date of Retirement (DOR)',
            'Phone Number', 'Email ID', 'S.B. Account Number', 'Bank Name',
            'Bank City Name', 'Branch Code', 'IFSC Code', 'Remarks', 'Password'
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Sample data row – dates in dd-mm-yyyy format
        sample_data = [
            'B.Tech', 'STF001', 'Teaching', 'Engineering',
            'Internal', 'John Doe', 'Professor', 'CSE',
            'CSE-Final', 'Engineering College', 'Mumbai', 'Mumbai',
            'PhD', '01-01-2020', '',  # DOR left blank
            '9876543210', 'john@example.com', '1234567890', 'State Bank of India',
            'Mumbai', '1234', 'SBIN0012345', 'Sample record', 'password123'
        ]

        for col_idx, value in enumerate(sample_data, start=1):
            ws.cell(row=2, column=col_idx, value=value)

        # Instructions (unchanged)
        notes_row = 4
        ws.cell(row=notes_row, column=1, value="Instructions:")
        ws.cell(row=notes_row, column=2, value="* Required fields")

        ws.cell(row=notes_row + 1, column=1, value="Date format:")
        ws.cell(row=notes_row + 1, column=2, value="dd-mm-yyyy (e.g., 15-01-2025)")

        ws.cell(row=notes_row + 2, column=1, value="Phone:")
        ws.cell(row=notes_row + 2, column=2, value="Must be 10 digits, unique")

        ws.cell(row=notes_row + 3, column=1, value="IFSC Code:")
        ws.cell(row=notes_row + 3, column=2, value="11 characters (e.g., SBIN0012345)")

        ws.cell(row=notes_row + 4, column=1, value="Note:")
        ws.cell(row=notes_row + 4, column=2, value="Remove the sample row before uploading your data")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Staff Template.xlsx'
        wb.save(response)
        return response

    except Exception as e:
        logger.error(f"Error generating sample template: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# -------------------------------------------------------------------------------------------------------------------------

@require_http_methods(["GET"])
@login_required
def staff_details(request, phone):
    
    try:
        staff = get_object_or_404(Staff, phone=phone)
        
        if SystemSetting.get_setting('log_detail_views', False):
            AuditLog.log(
                request=request,
                action='VIEW',
                obj=staff,
                object_repr=f"Staff details viewed: {staff.staff_id} - {staff.name}"
            )
        
        data = {
            'success': True,
            'staff': {
                'id': staff.id,
                'staff_id': staff.staff_id,
                'name': staff.name,
                'designation': staff.designation or '',
                'program': staff.program or '',
                'college': staff.college or '',
                'doj': staff.doj.strftime('%Y-%m-%d') if staff.doj else '',
                'dor': staff.dor.strftime('%Y-%m-%d') if staff.dor else '',
                'phone': staff.phone,
                'password': staff.password or '', 
                'email': staff.email or '',
                'district': staff.district or '',
                'place': staff.place or '',
                'qualification': staff.qualification or '',
                'bank_account': staff.bank_account or '',
                'bank_name': staff.bank_name or '',
                'bank_city': staff.bank_city or '',
                'branch_code': staff.branch_code or '',
                'ifsc_code': staff.ifsc_code or '',
                'remark': staff.remark or '',
                'staff_category': staff.staff_category or '',
                'dept_category': staff.dept_category or '',
                'examiner_type': staff.examiner_type or '',
                'branch': staff.branch or '',
                'branch_final': staff.branch_final or '',
                'created_at': staff.created_at.strftime('%Y-%m-%d %H:%M:%S') if staff.created_at else '',
                'updated_at': staff.updated_at.strftime('%Y-%m-%d %H:%M:%S') if staff.updated_at else '',
            }
        }
        return JsonResponse(data)
        
    except Exception as e:
        logger.error(f"Error getting staff details: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=400)