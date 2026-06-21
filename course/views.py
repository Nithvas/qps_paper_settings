from django.shortcuts import render, get_object_or_404
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from .models import Course
from syllabus.models import Syllabus
from pattern.models import QuestionPattern
from core.models import FieldReference, AuditLog, SystemSetting
import json
import logging

logger = logging.getLogger(__name__)

# ----------------------------------------------------------------------
# Course List – with per‑page, jump, sorting (Staff‑style)
# ----------------------------------------------------------------------

@login_required
def course_list(request):

    course_queryset = Course.objects.all()

    # --- Pagination settings ---
    ALLOWED_PER_PAGE = [25, 50, 75, 100, 500, 1000]
    default_per_page = SystemSetting.get_setting('course_items_per_page', 25)
    per_page = request.GET.get('per_page', default_per_page)
    try:
        per_page = int(per_page)
        if per_page not in ALLOWED_PER_PAGE:
            per_page = default_per_page
    except (ValueError, TypeError):
        per_page = default_per_page

    # --- Sorting ---
    sort_by = request.GET.get('sort', 'id')
    sort_dir = request.GET.get('dir', 'desc')

    valid_sort_fields = [
        'id', 'course_code', 'course_id', 'course_title', 'course_category',
        'program_type', 'degree', 'branch', 'branch_final', 'semester', 'part',
        'hours', 'credit', 'internal_mark', 'external_mark', 'total_mark',
        'examiner_type', 'created_at'
    ]

    if sort_by in valid_sort_fields:
        order = f"-{sort_by}" if sort_dir == 'desc' else sort_by
        course_queryset = course_queryset.order_by(order)
    else:
        course_queryset = course_queryset.order_by('-id')

    # --- Search ---
    search_query = request.GET.get('search', '')
    if search_query:
        course_queryset = course_queryset.filter(
            Q(course_code__icontains=search_query) |
            Q(course_id__icontains=search_query) |
            Q(course_title__icontains=search_query) |
            Q(course_category__icontains=search_query) |
            Q(program_type__icontains=search_query) |
            Q(degree__icontains=search_query) |
            Q(branch__icontains=search_query) |
            Q(branch_final__icontains=search_query) |
            Q(semester__icontains=search_query) |
            Q(part__icontains=search_query) |
            Q(examiner_type__icontains=search_query)
        )

    # --- Filters ---
    filters = {
        'program_type': request.GET.get('program_type', ''),
        'degree': request.GET.get('degree', ''),
        'branch': request.GET.get('branch', ''),
        'branch_final': request.GET.get('branch_final', ''),
        'course_category': request.GET.get('course_category', ''),
        'semester': request.GET.get('semester', ''),
        'part': request.GET.get('part', ''),
        'examiner_type': request.GET.get('examiner_type', ''),
    }

    for field, value in filters.items():
        if value:
            course_queryset = course_queryset.filter(**{field: value})

    # --- Filter options ---
    filter_options = {
        'program_types': FieldReference.get_options(Course, 'program_type'),
        'degrees': FieldReference.get_options(Course, 'degree'),
        'branches': FieldReference.get_options(Course, 'branch'),
        'branch_finals': FieldReference.get_options(Course, 'branch_final'),
        'course_categories': FieldReference.get_options(Course, 'course_category'),
        'semesters': FieldReference.get_options(Course, 'semester'),
        'parts': FieldReference.get_options(Course, 'part'),
        'examiner_types': FieldReference.get_options(Course, 'examiner_type'),
    }

    # --- Pagination ---
    paginator = Paginator(course_queryset, per_page)
    page_number = request.GET.get('page', 1)
    course_page = paginator.get_page(page_number)

    # Optional audit log
    if SystemSetting.get_setting('log_list_views', False):
        AuditLog.log(
            request=request,
            action='VIEW',
            obj=None,
            object_repr=f"Course list viewed - Page {page_number}",
            changes={
                'search': search_query,
                'filters': filters,
                'total_count': course_queryset.count()
            }
        )

    context = {
        'courses': course_page,
        'total_count': course_queryset.count(),
        'search': search_query,
        'filters': filters,
        'filter_options': filter_options,
        'current_sort': sort_by,
        'current_dir': sort_dir,
        'per_page': per_page,
        'allowed_per_page': ALLOWED_PER_PAGE,
    }
    return render(request, 'course/course_list.html', context)


# ----------------------------------------------------------------------
# Get field options (AJAX)
# ----------------------------------------------------------------------

@require_http_methods(["GET"])
@login_required
def get_field_options(request):

    field_name = request.GET.get('field')
    search_query = request.GET.get('search', '')

    if not field_name:
        return JsonResponse({'success': False, 'error': 'Field name required'}, status=400)

    field_mapping = {
        'program_type': 'program_type',
        'degree': 'degree',
        'branch': 'branch',
        'branch_final': 'branch_final',
        'course_category': 'course_category',
        'semester': 'semester',
        'part': 'part',
        'examiner_type': 'examiner_type',
        'filter_semester': 'semester',
        'filter_part': 'part',
        'filter_program_type': 'program_type',
        'filter_degree': 'degree',
        'filter_branch': 'branch',
        'filter_branch_final': 'branch_final',
        'filter_course_category': 'course_category',
        'filter_examiner_type': 'examiner_type',
    }

    model_field = field_mapping.get(field_name)
    if not model_field:
        return JsonResponse({'success': False, 'error': 'Invalid field'}, status=400)

    try:
        options = list(Course.objects.filter(
            **{f"{model_field}__isnull": False}
        ).exclude(
            **{f"{model_field}": ''}
        ).values_list(
            model_field, flat=True
        ).distinct().order_by(model_field))

        if search_query:
            options = [opt for opt in options if search_query.lower() in opt.lower()]

        return JsonResponse({'success': True, 'options': options, 'total_count': len(options)})
    except Exception as e:
        logger.error(f"Error getting field options: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ----------------------------------------------------------------------
# Add Course (GET form, POST save)
# ----------------------------------------------------------------------

@require_http_methods(["GET", "POST"])
@login_required
def course_add(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body) if request.content_type == 'application/json' else request.POST

            # Check uniqueness
            if Course.objects.filter(course_code=data.get('course_code')).exists():
                return JsonResponse({'success': False, 'error': 'Course code already exists'}, status=400)
            if Course.objects.filter(course_id=data.get('course_id')).exists():
                return JsonResponse({'success': False, 'error': 'Course ID already exists'}, status=400)

            course = Course()

            fields = [
                'program_type', 'degree', 'branch', 'branch_final',
                'course_code', 'course_id', 'course_category', 'course_title',
                'semester', 'part', 'hours', 'credit',
                'internal_mark', 'external_mark', 'total_mark',
                'examiner_type', 'remark'
            ]

            changes = {}
            for f in fields:
                val = data.get(f)
                if val and str(val).strip():
                    setattr(course, f, str(val).strip())
                    changes[f] = str(val).strip()
                else:
                    setattr(course, f, None)

            # Auto‑calculate total mark
            try:
                internal = int(course.internal_mark) if course.internal_mark else 0
                external = int(course.external_mark) if course.external_mark else 0
                if not course.total_mark:
                    course.total_mark = str(internal + external)
                    changes['total_mark'] = course.total_mark
            except ValueError:
                pass

            course.full_clean()
            course.save()

            AuditLog.log(
                request=request,
                action='CREATE',
                obj=course,
                object_repr=f"{course.course_code} - {course.course_title}",
                changes=changes
            )

            return JsonResponse({'success': True, 'message': 'Course added successfully', 'id': course.id})

        except ValidationError as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        except Exception as e:
            logger.error(f"Error adding course: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

    # GET – return form options (for drawer)
    context = {
        'field_options': {
            'program_types': FieldReference.get_options(Course, 'program_type'),
            'degrees': FieldReference.get_options(Course, 'degree'),
            'branches': FieldReference.get_options(Course, 'branch'),
            'branch_finals': FieldReference.get_options(Course, 'branch_final'),
            'course_categories': FieldReference.get_options(Course, 'course_category'),
            'semesters': FieldReference.get_options(Course, 'semester'),
            'parts': FieldReference.get_options(Course, 'part'),
            'examiner_types': FieldReference.get_options(Course, 'examiner_type'),
            'patterns': QuestionPattern.objects.filter(is_active=True).values_list('pattern_code', flat=True),
        }
    }
    return render(request, 'course/course_form.html', context)


# ----------------------------------------------------------------------
# Edit Course
# ----------------------------------------------------------------------

@require_http_methods(["GET", "POST"])
@login_required
def course_edit(request, course_code):
    course = get_object_or_404(Course, course_code=course_code)

    if request.method == "GET":
        data = {
            'success': True,
            'course': {
                'id': course.id,
                'program_type': course.program_type or '',
                'degree': course.degree or '',
                'branch': course.branch or '',
                'branch_final': course.branch_final or '',
                'course_code': course.course_code,
                'course_id': course.course_id or '',
                'course_category': course.course_category or '',
                'course_title': course.course_title,
                'semester': course.semester or '',
                'part': course.part or '',
                'hours': course.hours or '',
                'credit': course.credit or '',
                'internal_mark': course.internal_mark or '',
                'external_mark': course.external_mark or '',
                'total_mark': course.total_mark or '',
                'examiner_type': course.examiner_type or '',
                'remark': course.remark or '',
                'pattern': course.pattern.pattern_code if course.pattern else '',
            }
        }
        return JsonResponse(data)

    # POST update
    try:
        old_values = {f: getattr(course, f) for f in [
            'program_type', 'degree', 'branch', 'branch_final',
            'course_code', 'course_id', 'course_category', 'course_title',
            'semester', 'part', 'hours', 'credit',
            'internal_mark', 'external_mark', 'total_mark',
            'examiner_type', 'remark'
        ]}

        changes = {}
        for f in old_values.keys():
            new_val = request.POST.get(f)
            if new_val and str(new_val).strip():
                new_val = str(new_val).strip()
                setattr(course, f, new_val)
                if str(old_values[f]) != new_val:
                    changes[f] = {'old': str(old_values[f]), 'new': new_val}
            else:
                setattr(course, f, None)
                if old_values[f] is not None and old_values[f] != '':
                    changes[f] = {'old': str(old_values[f]), 'new': None}

        # Auto‑calculate total
        try:
            internal = int(course.internal_mark) if course.internal_mark else 0
            external = int(course.external_mark) if course.external_mark else 0
            new_total = str(internal + external)
            if course.total_mark != new_total:
                course.total_mark = new_total
                changes['total_mark'] = {'old': old_values.get('total_mark', ''), 'new': new_total}
        except ValueError:
            pass

        # Handle pattern (Template ID)
        pattern_code = request.POST.get('pattern')
        if pattern_code:
            try:
                pattern = QuestionPattern.objects.get(pattern_code=pattern_code, is_active=True)
                if course.pattern != pattern:
                    course.pattern = pattern
                    changes['pattern'] = {'old': course.pattern.pattern_code if course.pattern else '', 'new': pattern_code}
            except QuestionPattern.DoesNotExist:
                pass
        else:
            if course.pattern:
                course.pattern = None
                changes['pattern'] = {'old': old_values.get('pattern', ''), 'new': None}

        # Check uniqueness if changed
        new_code = request.POST.get('course_code')
        if new_code and new_code.strip() and new_code.strip() != old_values['course_code']:
            if Course.objects.filter(course_code=new_code.strip()).exists():
                return JsonResponse({'success': False, 'error': 'Course code already exists'}, status=400)
            course.course_code = new_code.strip()

        new_id = request.POST.get('course_id')
        if new_id and new_id.strip() and new_id.strip() != old_values['course_id']:
            if Course.objects.filter(course_id=new_id.strip()).exists():
                return JsonResponse({'success': False, 'error': 'Course ID already exists'}, status=400)
            course.course_id = new_id.strip()

        course.full_clean()
        course.save()

        if changes:
            AuditLog.log(
                request=request,
                action='UPDATE',
                obj=course,
                object_repr=f"{course.course_code} - {course.course_title}",
                changes=changes
            )

        return JsonResponse({'success': True, 'message': 'Course updated successfully'})

    except ValidationError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        logger.error(f"Error editing course: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ----------------------------------------------------------------------
# Delete Course
# ----------------------------------------------------------------------

@require_http_methods(["DELETE", "POST"])
@login_required
def course_delete(request, course_code):
    try:
        course = get_object_or_404(Course, course_code=course_code)
        course_info = f"{course.course_code} - {course.course_title}"
        course_data = {
            'course_code': course.course_code,
            'course_id': course.course_id,
            'course_title': course.course_title,
            'program_type': course.program_type,
            'degree': course.degree,
            'branch': course.branch,
            'semester': course.semester
        }

        if SystemSetting.get_setting('soft_delete_course', False):
            course.is_active = False
            course.save()
            AuditLog.log(
                request=request,
                action='UPDATE',
                obj=course,
                object_repr=f"Soft deleted: {course_info}",
                changes=course_data
            )
            return JsonResponse({'success': True, 'message': 'Course soft deleted'})
        else:
            course.delete()
            AuditLog.objects.create(
                action='DELETE',
                app_label=Course._meta.app_label,
                model_name=Course.__name__,
                object_id=str(course.id),
                object_repr=course_info,
                changes=course_data,
                ip_address=AuditLog._get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                request_path=request.path,
                user_id=request.user.id,
                user_name=request.user.username
            )
            return JsonResponse({'success': True, 'message': 'Course deleted permanently'})

    except Exception as e:
        logger.error(f"Error deleting course: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ----------------------------------------------------------------------
# Save new field option
# ----------------------------------------------------------------------

@require_http_methods(["POST"])
@login_required
def save_field_option(request):
    try:
        data = json.loads(request.body)
        field_name = data.get('field_name')
        value = data.get('value')

        valid_fields = [
            'program_type', 'degree', 'branch', 'branch_final', 'course_category',
            'semester', 'part', 'examiner_type'
        ]

        if field_name not in valid_fields or not value:
            return JsonResponse({'success': False, 'error': 'Invalid field or value'}, status=400)

        obj, created = FieldReference.add_option(
            model=Course,
            field_name=field_name,
            value=value,
            created_by=request.user.username
        )

        return JsonResponse({'success': True, 'created': created, 'message': 'Option saved'})
    except Exception as e:
        logger.error(f"Error saving field option: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ----------------------------------------------------------------------
# Upload Courses from Excel
# ----------------------------------------------------------------------

@require_http_methods(["POST"])
@login_required
def course_upload(request):
    
    if not request.FILES.get('file'):
        return JsonResponse({'success': False, 'error': 'No file provided'}, status=400)

    try:
        file = request.FILES['file']
        if not file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({'success': False, 'error': 'Invalid file format. Please upload Excel file.'}, status=400)

        wb = load_workbook(file)
        ws = wb.active

        success_count = 0
        update_count = 0
        error_count = 0
        errors = []
        added_options_count = 0

        def safe_str(val, default=''):
            return str(val).strip() if val else default

        def safe_int(val, default=0):
            try:
                return int(float(val)) if val else default
            except (ValueError, TypeError):
                return default

        new_options = {
            'program_type': set(),
            'degree': set(),
            'branch': set(),
            'branch_final': set(),
            'course_category': set(),
            'semester': set(),
            'part': set(),
            'examiner_type': set(),
        }

        REQUIRED_FIELDS = [
            'program_type', 'degree', 'branch', 'branch_final',
            'course_category', 'semester',
        ]

        # Helper to ensure a Syllabus exists – now uses Course.id via the instance
        def ensure_syllabus(course_instance):
            Syllabus.objects.get_or_create(course=course_instance)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            if not row[6] or not row[7]:
                err_msg = f"Row {row_idx}: Course Code (col 7) and Title (col 8) are required"
                errors.append(err_msg)
                logger.error(err_msg)
                error_count += 1
                continue

            program_type = safe_str(row[0] if len(row) > 0 else None)
            degree = safe_str(row[1] if len(row) > 1 else None)
            branch = safe_str(row[2] if len(row) > 2 else None)
            branch_final = safe_str(row[3] if len(row) > 3 else None)
            course_id = safe_str(row[4] if len(row) > 4 else None)
            course_category = safe_str(row[5] if len(row) > 5 else None)
            course_code = safe_str(row[6] if len(row) > 6 else None)
            course_title = safe_str(row[7] if len(row) > 7 else None)
            semester = safe_str(row[8] if len(row) > 8 else None)
            part = safe_str(row[9] if len(row) > 9 else None)
            hours = safe_int(row[10] if len(row) > 10 else None)
            credit = safe_str(row[11] if len(row) > 11 else None)
            internal_mark = safe_int(row[12] if len(row) > 12 else None)
            external_mark = safe_int(row[13] if len(row) > 13 else None)
            total_mark = safe_int(row[14] if len(row) > 14 else None,
                                  safe_int(row[12] if len(row) > 12 else None) + safe_int(row[13] if len(row) > 13 else None))
            examiner_type = safe_str(row[15] if len(row) > 15 else None)
            pattern_code = safe_str(row[16] if len(row) > 16 else None)
            remark = safe_str(row[17] if len(row) > 17 else None)

            missing_fields = []
            for field in REQUIRED_FIELDS:
                value = locals().get(field)
                if not value:
                    missing_fields.append(field)

            if missing_fields:
                err_msg = f"Row {row_idx}: Missing required fields: {', '.join(missing_fields)}"
                errors.append(err_msg)
                logger.error(err_msg)
                error_count += 1
                continue

            col_map = [0, 1, 2, 3, 5, 8, 9, 15]
            for idx, field in enumerate(['program_type','degree','branch','branch_final','course_category','semester','part','examiner_type']):
                val = safe_str(row[col_map[idx]] if len(row) > col_map[idx] else None)
                if val:
                    new_options[field].add(val)

            course_data = {
                'program_type': program_type,
                'degree': degree,
                'branch': branch,
                'branch_final': branch_final,
                'course_id': course_id,
                'course_category': course_category,
                'course_code': course_code,
                'course_title': course_title,
                'semester': semester,
                'part': part,
                'hours': hours,
                'credit': credit,
                'internal_mark': internal_mark,
                'external_mark': external_mark,
                'total_mark': total_mark,
                'examiner_type': examiner_type,
                'remark': remark,
            }

            if pattern_code:
                try:
                    pattern = QuestionPattern.objects.get(pattern_code=pattern_code, is_active=True)
                    course_data['pattern'] = pattern
                except QuestionPattern.DoesNotExist:
                    err_msg = f"Row {row_idx}: Pattern '{pattern_code}' not found"
                    errors.append(err_msg)
                    logger.error(err_msg)
                    error_count += 1
                    continue

            existing = Course.objects.filter(course_code=course_code).first()
            try:
                if existing:
                    for key, val in course_data.items():
                        if key != 'course_code' and val is not None:
                            setattr(existing, key, val)
                    existing.full_clean()
                    existing.save()
                    ensure_syllabus(existing)  
                    update_count += 1
                else:
                    new_course = Course(**course_data)
                    new_course.full_clean()
                    new_course.save()
                    ensure_syllabus(new_course) 
                    success_count += 1

            except ValidationError as ve:
                if hasattr(ve, 'message_dict'):
                    for field, field_errors in ve.message_dict.items():
                        err_msg = f"Row {row_idx}: Field '{field}' - {', '.join(field_errors)}"
                        errors.append(err_msg)
                        logger.error(err_msg)
                else:
                    err_msg = f"Row {row_idx}: {ve.messages[0] if ve.messages else str(ve)}"
                    errors.append(err_msg)
                    logger.error(err_msg)
                error_count += 1

            except Exception as e:
                err_msg = f"Row {row_idx}: {str(e)}"
                errors.append(err_msg)
                logger.error(err_msg)
                error_count += 1

        for field, values in new_options.items():
            for val in values:
                if val:
                    obj, created = FieldReference.add_option(
                        model=Course,
                        field_name=field,
                        value=val,
                        created_by=request.user.username
                    )
                    if created:
                        added_options_count += 1

        AuditLog.objects.create(
            action='UPLOAD',
            app_label=Course._meta.app_label,
            model_name=Course.__name__,
            object_repr=f"Bulk upload: {success_count} new, {update_count} updated",
            changes={
                'filename': file.name,
                'success_count': success_count,
                'update_count': update_count,
                'error_count': error_count,
                'new_options_added': added_options_count,
            },
            ip_address=AuditLog._get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            request_path=request.path,
            user_id=request.user.id,
            user_name=request.user.username
        )

        msg = f"Processed: {success_count} added, {update_count} updated, {error_count} errors."
        if error_count:
            messages.warning(request, msg)
        else:
            messages.success(request, msg)

        return JsonResponse({
            'success': True,
            'success_count': success_count,
            'update_count': update_count,
            'error_count': error_count,
            'errors': errors[:20],
            'new_options_added': added_options_count,
        })

    except Exception as e:
        logger.error(f"Error in course upload: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ----------------------------------------------------------------------
# Export Courses to Excel 
# ----------------------------------------------------------------------

@login_required
def course_export(request):

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Course Data"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        headers = [
            'Program', 'Degree', 'Branch', 'Branch Final',
            'Course ID', 'Course Category', 'Course Code', 'Course Title',
            'Semester', 'Part', 'Hours', 'Credit',
            'Internal Mark', 'External Mark', 'Total Mark',
            'Examiner Type', 'Template ID'
        ]

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        courses = Course.objects.all().order_by('id')
        for row_idx, course in enumerate(courses, start=2):
            ws.cell(row=row_idx, column=1, value=course.program_type or '')
            ws.cell(row=row_idx, column=2, value=course.degree or '')
            ws.cell(row=row_idx, column=3, value=course.branch or '')
            ws.cell(row=row_idx, column=4, value=course.branch_final or '')
            ws.cell(row=row_idx, column=5, value=course.course_id or '')
            ws.cell(row=row_idx, column=6, value=course.course_category or '')
            ws.cell(row=row_idx, column=7, value=course.course_code or '')
            ws.cell(row=row_idx, column=8, value=course.course_title or '')
            ws.cell(row=row_idx, column=9, value=course.semester or '')
            ws.cell(row=row_idx, column=10, value=course.part or '')
            ws.cell(row=row_idx, column=11, value=course.hours or '')
            ws.cell(row=row_idx, column=12, value=course.credit or '')
            ws.cell(row=row_idx, column=13, value=course.internal_mark or '')
            ws.cell(row=row_idx, column=14, value=course.external_mark or '')
            ws.cell(row=row_idx, column=15, value=course.total_mark or '')
            ws.cell(row=row_idx, column=16, value=course.examiner_type or '')
            ws.cell(row=row_idx, column=17, value=course.pattern.pattern_code if course.pattern else '')

        # Auto‑adjust widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # Audit log
        AuditLog.objects.create(
            action='EXPORT',
            app_label=Course._meta.app_label,
            model_name=Course.__name__,
            object_repr=f"Course export - {courses.count()} records",
            changes={'total_records': courses.count(), 'file_format': 'xlsx'},
            ip_address=AuditLog._get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            request_path=request.path,
            user_id=request.user.id,
            user_name=request.user.username
        )

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Course Data.xlsx'
        wb.save(response)
        return response

    except Exception as e:
        logger.error(f"Error exporting courses: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ----------------------------------------------------------------------
# Download Sample Template 
# ----------------------------------------------------------------------

@login_required
def course_sample(request):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Course Template"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        headers = [
            'Program', 'Degree', 'Branch', 'Branch Final',
            'Course ID', 'Course Category', 'Course Code*', 'Course Title*',
            'Semester', 'Part', 'Hours', 'Credit',
            'Internal Mark', 'External Mark', 'Total Mark',
            'Examiner Type', 'Template ID'
        ]

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Sample row
        sample = [
            'Regular', 'B.Tech', 'Computer Science', 'CSE Final',
            'CS101', 'Core', 'CS101A', 'Programming Basics',
            '1', '1', '4', '3.0',
            '40', '60', '100',
            'Internal', 'PAT-001'
        ]
        for col_idx, val in enumerate(sample, start=1):
            ws.cell(row=2, column=col_idx, value=val)

        # Notes
        notes = [
            ('* Required fields', ''),
            ('Course Code must be unique', ''),
            ('Course ID must be unique', ''),
            ('Template ID is the pattern_code from QuestionPattern', ''),
            ('Remove sample row before uploading', ''),
        ]
        for i, (note, _) in enumerate(notes, start=4):
            ws.cell(row=i, column=1, value=note)

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Course Template.xlsx'
        wb.save(response)
        return response

    except Exception as e:
        logger.error(f"Error generating sample: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ----------------------------------------------------------------------
# Course Details (AJAX)
# ----------------------------------------------------------------------

@require_http_methods(["GET"])
@login_required
def course_details(request, course_code):

    try:
        course = get_object_or_404(Course, course_code=course_code)
        data = {
            'success': True,
            'course': {
                'id': course.id,
                'program_type': course.program_type or '',
                'degree': course.degree or '',
                'branch': course.branch or '',
                'branch_final': course.branch_final or '',
                'course_code': course.course_code,
                'course_id': course.course_id or '',
                'course_category': course.course_category or '',
                'course_title': course.course_title,
                'semester': course.semester or '',
                'part': course.part or '',
                'hours': course.hours or '',
                'credit': course.credit or '',
                'internal_mark': course.internal_mark or '',
                'external_mark': course.external_mark or '',
                'total_mark': course.total_mark or '',
                'examiner_type': course.examiner_type or '',
                'remark': course.remark or '',
                'pattern_code': course.pattern.pattern_code if course.pattern else '',
                'created_at': course.created_at.strftime('%Y-%m-%d %H:%M:%S') if course.created_at else '',
                'updated_at': course.updated_at.strftime('%Y-%m-%d %H:%M:%S') if course.updated_at else '',
            }
        }
        return JsonResponse(data)
    except Exception as e:
        logger.error(f"Error getting course details: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=400)