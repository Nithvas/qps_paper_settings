import openpyxl
from django.http import HttpResponse
from ..models import Course

# -------------------------------------------------------------------
# FIELD ORDER FOR EXCEL FILES
# -------------------------------------------------------------------

COURSE_EXCEL_FIELDS = [
    "course_code",
    "course_title",
    "semester",
    "course_id",
    "program",
    "external_mark",
    "examiner",
]

# -------------------------------------------------------------------
# SAMPLE FILE
# -------------------------------------------------------------------

def download_sample_course_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(COURSE_EXCEL_FIELDS)
    ws.append(["23MCA1CC1", "Data Structures", "5", "MCA", "UG","70", "Internal"])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Course Template.xlsx"'
    wb.save(response)
    return response

# -------------------------------------------------------------------
# EXPORT (MODEL → EXCEL)
# -------------------------------------------------------------------

def export_course_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(COURSE_EXCEL_FIELDS)
    for c in Course.objects.all():
        ws.append([
            c.course_code,
            c.course_title,
            c.semester,
            c.course_id,
            c.program,
            c.external_mark,
            c.examiner,
        ])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Course Data.xlsx"'
    wb.save(response)
    return response

# -------------------------------------------------------------------
# UPLOAD 
# -------------------------------------------------------------------

def upload_course_excel(file):

    wb = openpyxl.load_workbook(file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else "" for h in rows[0]]
    expected = {f.lower() for f in COURSE_EXCEL_FIELDS}
    if not expected.issubset({h.lower() for h in headers}):
        raise ValueError("Excel headers do not match required fields. Required: " + ", ".join(COURSE_EXCEL_FIELDS))
    
    for row in rows[1:]:
        data = dict(zip(headers, row))
        if not any(data.values()):
            continue
        Course.objects.create(
            course_code=data.get("course_code"),
            course_title=data.get("course_title"),
            semester=data.get("semester"),
            course_id=data.get("course_id"),
            program=data.get("program"),
            external_mark=data.get("external_mark"),
            examiner=data.get("examiner"),
        )